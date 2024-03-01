# from openpyxl import load_workbook
# from PIL import Image, ImageDraw, ImageFont

# # Cargar la base de datos desde el archivo Excel
# excel_file_path = r"C:/Users/SEBASTIAN/OneDrive/Escritorio/Registro de Bonos .xlsx"
# workbook = load_workbook(excel_file_path)
# sheet = workbook.active

# # Ruta de la imagen del bono
# imagen_bono_path = r"C:/Users/SEBASTIAN/OneDrive/Escritorio/-135.jpg"
# output_path = r"C:/Users/SEBASTIAN/OneDrive/Escritorio/bono final"


# # Obtener los títulos de las columnas desde la primera fila
# headers = [col[0].value for col in sheet.iter_cols()]

# # Iterar sobre las filas de la hoja de cálculo
# for row in sheet.iter_rows(min_row=2, values_only=True):
#     nombre = row[headers.index('NOMBRES ')]
#     apellido = row[headers.index('APELLIDOS')]
#     cedula = row[headers.index('C.C')]
#     valor = row[headers.index('VALOR DEL BONO')]
#     numero_bono = row[headers.index('# BONO')]


#     # Abrir la imagen del bono
#     bono = Image.open(imagen_bono_path)

#     # Crear un objeto ImageDraw para agregar texto a la imagen
#     draw = ImageDraw.Draw(bono)

#     # Definir la fuente y el tamaño del texto
#     font = ImageFont.load_default()  # Puedes cambiar esto según tus preferencias

#     # Agregar la información al bono
#     draw.text((197, 69), f"{nombre}", font=font, fill=(255, 255, 255))
#     draw.text((218, 63), f"{apellido}", font=font, fill=(255, 255, 255))
#     draw.text((308, 57), f"{cedula}", font=font, fill=(255, 255, 255))
#     draw.text((253, 59), f"{valor}", font=font, fill=(255, 255, 255))
#     draw.text((179, 87), f"{numero_bono}", font=font, fill=(255, 255, 255))

#     # Guardar la nueva imagen con la información
#     output_file_path = output_path.format(nombre, cedula) + ".jpg"  # o el formato que desees
#     try:
#      bono.save(output_file_path)
#     except Exception as e:
#         print(f"Error al guardar la imagen: {e}")
#     print(output_file_path)
#     from PIL import Image


# # Cerrar el archivo Excel después de usarlo
# workbook.close()
# import os

# # ...

# # Verificar permisos de escritura en la carpeta de salida
# output_folder = os.path.dirname(output_path)
# if os.access(output_folder, os.W_OK):
#     print(f"Tienes permisos de escritura en la carpeta: {output_folder}")
# else:
#     print(f"No tienes permisos de escritura en la carpeta: {output_folder}")
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont
import os

# Cargar la base de datos desde el archivo Excel
excel_file_path = r"C:/Users/SEBASTIAN/OneDrive/Escritorio/Registro de Bonos .xlsx"
workbook = load_workbook(excel_file_path)
sheet = workbook.active

# Ruta de la imagen del bono
imagen_bono_path = r"C:/Users/SEBASTIAN/OneDrive/Escritorio/-135.jpg"
output_path = "C:\\Users\\SEBASTIAN\\OneDrive\\Escritorio\\bono"



# Obtener los títulos de las columnas desde la primera fila
headers = [col[0].value for col in sheet.iter_cols()]

# Iterar sobre las filas de la hoja de cálculo
for row in sheet.iter_rows(min_row=2, values_only=True):
    nombre = row[headers.index('NOMBRES ')]
    apellido = row[headers.index('APELLIDOS')]
    cedula = row[headers.index('C.C')]
    valor = row[headers.index('VALOR DEL BONO')]
    numero_bono = row[headers.index('# BONO')]

    # Abrir la imagen del bono
    bono = Image.open(imagen_bono_path)

    # Crear un objeto ImageDraw para agregar texto a la imagen
    draw = ImageDraw.Draw(bono)

    # Tamaño y tipo de letra
    font_size = 40  # Ajusta este valor según tus preferencias
    font_type = "C:\\Users\\SEBASTIAN\\OneDrive\\Escritorio\\Forum-Regular.otf"  # Reemplaza con la ruta real de tu fuente .ttf

    # Cargar la fuente con el tamaño deseado
    font = ImageFont.truetype(font_type, font_size)

    # Agregar la información al bono
    draw.text((458,323 ), f"{nombre}", font=font, fill=(0, 0, 0))
    draw.text((657,323 ), f"{apellido}", font=font, fill=(0, 0, 0))
    draw.text((530,392 ), f"{cedula}", font=font, fill=(0, 0, 0))
    draw.text((337,445 ), f"{valor}", font=font, fill=(0, 0, 0))
    draw.text((1308,112 ), f"{numero_bono}", font=font, fill=(0, 0, 0))

    # Guardar la nueva imagen con la información
    output_file_path = os.path.join(output_path, f"{nombre}_{cedula}.jpg")
    
    try:
        bono.save(output_file_path)
        print(f"Imagen guardada con éxito en: {output_file_path}")
    except Exception as e:
        print(f"Error al guardar la imagen: {e}")

# Cerrar el archivo Excel después de usarlo
workbook.close()
